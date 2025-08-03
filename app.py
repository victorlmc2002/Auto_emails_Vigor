from openpyxl import load_workbook
from devedores import Devedores
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from pathlib import Path
import re
from itertools import groupby
import tkinter as tk
from tkinter import ttk, messagebox
from email.mime.image import MIMEImage

def ler_txt(path, lista):
    with open(path, "r", encoding="utf-8") as arquivo:
        for linha in arquivo:
            linha = linha.strip()
            if linha: 
                partes = linha.split(". ", 1)
                if len(partes) == 2:
                    lista.append(partes[1])
                else:
                    lista.append(linha)
            else:
                lista.append("")

# Configurações
MES_ANTERIOR = "Maio"
MES = "06 - Junho"
EMAIL_FROM = 'vigor.imob@gmail.com'
#EMAIL_PASSWORD = 'uryt lswm ptkb kuwq'
EMAIL_PASSWORD = 'qzls bxhy uiei mtiv'
EMAIL_TO_PROP = 'victorlmc2002@gmail.com'
EMAIL_TO_INQUILINO = 'suporte@vigornegocios.com.br'

QTD = 57  # Quantidade de imóveis a serem processados

# Lista de endereços
ENDERECOS = []
ler_txt("C:/Users/victo/Desktop/auto/Auto_emails_Vigor/enderecos.txt", ENDERECOS)

# Lista de nomes
NOMES_INQUELINOS = []
ler_txt("C:/Users/victo/Desktop/auto/Auto_emails_Vigor/nomes_inquelinos.txt", NOMES_INQUELINOS)

# Lista de emails
EMAILS_INQUELINOS = []
ler_txt("C:/Users/victo/Desktop/auto/Auto_emails_Vigor/emails_inquelinos.txt", EMAILS_INQUELINOS)

# Verifica se as listas têm o mesmo tamanho
if not (len(ENDERECOS) == len(NOMES_INQUELINOS) == len(EMAILS_INQUELINOS)):
    raise ValueError("As listas de endereços, nomes e emails devem ter o mesmo tamanho.")

# Caminhos das pastas
BASE_PATH = Path('C:/Users/victo/Desktop/teste/boletos')
PASTA_BOLETOS = BASE_PATH / MES / 'Boletos'
#print(PASTA_BOLETOS)
PASTA_CONDOMINIO = BASE_PATH / MES / 'Taxa de Condomínio'
PASTA_REPASSES = BASE_PATH / MES / 'Repasses'
ARQ_EXCEL = BASE_PATH / 'Planilha nova 2025.xlsx'

class EmailInterface:
    def __init__(self, devedores):
        self.devedores = devedores
        self.selected_items = set()
        self.root = tk.Tk()
        self.root.title("Envio de Emails - Sistema de Boletos")
        self.root.geometry("800x600")
        
        self.create_widgets()
        
    def create_widgets(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Título
        ttk.Label(main_frame, text="Selecione os imóveis para envio de emails", 
                 font=('Helvetica', 14, 'bold')).pack(pady=10)
        
        # Frame de seleção
        selection_frame = ttk.Frame(main_frame)
        selection_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Lista de imóveis com scrollbar
        self.tree = ttk.Treeview(selection_frame, columns=('Nome', 'Valor'), show='headings')
        self.tree.heading('Nome', text='Proprietário/Inquilino')
        self.tree.heading('Valor', text='Endereço')
        
        vsb = ttk.Scrollbar(selection_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Preencher a lista
        for i, devedor in enumerate(self.devedores, 1):
            self.tree.insert("", "end", values=(f"{i} - {devedor._nome_inquilino}", f"{devedor._endereco}"), tags=(str(i),))
        
        # Configurar tags para seleção
        self.tree.tag_configure('selected', background='lightblue')
        
        # Opções de envio
        options_frame = ttk.Frame(main_frame)
        options_frame.pack(fill=tk.X, pady=10)
        
        self.recipient_var = tk.StringVar(value="inquilino")
        ttk.Radiobutton(options_frame, text="Enviar para Inquilino", variable=self.recipient_var, 
                       value="inquilino").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(options_frame, text="Enviar para Proprietário", variable=self.recipient_var, 
                       value="proprietario").pack(side=tk.LEFT, padx=10)
        
        # Botões
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(button_frame, text="Selecionar Todos", command=self.select_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Desmarcar Todos", command=self.deselect_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Enviar Selecionados", command=self.send_selected).pack(side=tk.RIGHT, padx=5)
        
        # Configurar seleção múltipla
        self.tree.bind("<Button-1>", self.on_click)
    
    def on_click(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            tags = self.tree.item(item, "tags")
            if tags:
                idx = int(tags[0])
                if idx in self.selected_items:
                    self.selected_items.remove(idx)
                    self.tree.item(item, tags=(tags[0],))
                else:
                    self.selected_items.add(idx)
                    self.tree.item(item, tags=(tags[0], 'selected'))
    
    def select_all(self):
        self.selected_items = set(range(1, len(self.devedores)+1))
        for item in self.tree.get_children():
            tags = self.tree.item(item, "tags")
            if tags:
                self.tree.item(item, tags=(tags[0], 'selected'))
    
    def deselect_all(self):
        self.selected_items = set()
        for item in self.tree.get_children():
            tags = self.tree.item(item, "tags")
            if tags:
                self.tree.item(item, tags=(tags[0],))
    
    def send_selected(self):
        if not self.selected_items:
            messagebox.showwarning("Nenhum selecionado", "Por favor, selecione pelo menos um imóvel.")
            return
        
        recipient_type = self.recipient_var.get()
        confirm = messagebox.askyesno(
            "Confirmar envio",
            f"Você está prestes a enviar emails para {len(self.selected_items)} imóveis ({recipient_type}s).\nDeseja continuar?"
        )
        
        if not confirm:
            return
        
        self.root.config(cursor="watch")
        self.root.update()
        
        try:
            for idx in self.selected_items:
                devedor = self.devedores[idx-1]
                if recipient_type == "inquilino":
                    enviar_email_inquilino(devedor)
                else:
                    enviar_email_proprietario(devedor)
            
            messagebox.showinfo("Sucesso", "Emails enviados com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao enviar os emails:\n{str(e)}")
        finally:
            self.root.config(cursor="")
    
    def run(self):
        self.root.mainloop()

def extrair_numero(nome_arquivo):
    """Extrai número do nome do arquivo para ordenação."""
    match = re.search(r'\d+', nome_arquivo.stem)
    return int(match.group()) if match else float('inf')

def processar_pasta(pasta_path):
    """Processa arquivos PDF em uma pasta, agrupando por número."""
    arquivos = sorted(pasta_path.glob('*.pdf'), key=extrair_numero)
    
    grupos = []
    for key, group in groupby(arquivos, key=extrair_numero):
        grupo = list(group)
        grupos.append([str(p) for p in grupo] if len(grupo) > 1 else str(grupo[0]))
    
    return grupos

def ler_planilha(excel, n):
    """Lê dados da planilha e cria lista de devedores."""
    wb = load_workbook(excel, data_only=True)
    planilha = wb[MES]
    
    devedores = []
    imovel = 0
    x = 5
    pdfs = processar_pasta(PASTA_BOLETOS)
    conds = processar_pasta(PASTA_CONDOMINIO)
    repasses = processar_pasta(PASTA_REPASSES)
    while imovel < n:
        for y in range(2, 17, 5):
            if imovel >= n:
                break
            #print(x, y)
            #print(imovel)
            dados = extrair_dados_linha(planilha, x, y)
            
            devedor = criar_devedor(
                ENDERECOS[imovel], NOMES_INQUELINOS[imovel], EMAILS_INQUELINOS[imovel], dados, 
                pdfs[imovel] if imovel < len(pdfs) else None,
                conds[imovel] if imovel < len(conds) else None,
                repasses[imovel] if imovel < len(repasses) else None
            )
            print(devedor)
            devedores.append(devedor)
            imovel += 1
        x += 17
    wb.close()
    return devedores

def extrair_dados_linha(planilha, x, y):
    """Extrai dados de uma linha específica da planilha."""
    dados = {
        'aluguel': None,
        'iptu': None,
        'cond': None,
        'valor': None,
        'cotas_extras': [],
        'taxa': None,
        'tarifa': None,
        'repasse': None
    }
    index = 0
    
    while planilha.cell(row=x, column=y).value != "Receita":
        celula = planilha.cell(row=x, column=y)
        valor = planilha.cell(row=x, column=y+1).value
        #print(celula.value, valor)
        if celula.value == "Aluguel":
            dados['aluguel'] = valor
        elif celula.value == "Iptu cota":
            dados['iptu'] = valor
        elif celula.value == "Taxa de Condomínio":
            dados['cond'] = valor
        elif celula.value == "Valor Boleto":
            dados['valor'] = valor
        elif celula.value is None or "BLT" in celula.value:
            index = 1
        elif celula.value is not None and index == 0 and (float(valor) > 0):
            dados['cotas_extras'].append({celula.value: valor})
        
        x += 1
    
    dados['taxa'] = planilha.cell(row=x+2, column=y+1).value
    dados['tarifa'] = planilha.cell(row=x+2, column=y+3).value
    dados['repasse'] = planilha.cell(row=x+3, column=y+1).value
    
    return dados

def criar_devedor(endereco, nome, email, dados, pdf, cond, repasse):
    """Cria um objeto Devedor com os dados fornecidos."""
    return Devedores(
        endereco, nome, email, dados['valor'], dados['aluguel'], dados['iptu'], dados['cond'], 
        dados['cotas_extras'], dados['repasse'], dados['taxa'], dados['tarifa'],
        pdf, None if dados['cond'] is not None else cond, repasse
    )

def formatar_texto_inquilino(devedor):
    """Formata texto do e-mail para inquilinos."""
    partes = [f"Aluguel: R$ {devedor._aluguel:.2f}"]
    
    if not isinstance(devedor._iptu, str) and devedor._iptu is not None and devedor._iptu > 0:
        partes.append(f"IPTU: R$ {devedor._iptu:.2f}")
    if not isinstance(devedor._cond, str) and devedor._cond is not None and devedor._cond > 0:
        partes.append(f"Taxa de condomínio: R$ {devedor._cond:.2f}")
    if devedor._cotas_extras:
        partes.append("<br><b>Descontos/Reembolsos:<br></b>")
        for cota in devedor._cotas_extras:
            for chave, valor in cota.items():
                if valor > 0:
                    partes.append(f"{chave}: R$ {valor:.2f}")
    
    partes.append(f"<br><b>Valor do Boleto: R$ {devedor._valor:.2f}<br><br>Atenciosamente,<br>Vigor Negócios<br></b>")
    return "<br>".join(partes)

def formatar_texto_proprietario(devedor):
    """Formata texto do e-mail para proprietários."""
    partes = [f"Receita: R$ {devedor._valor:.2f}", "<br><b>Reembolso de Despesas:</b><br>"]
    
    if not isinstance(devedor._iptu, str) and devedor._iptu is not None and devedor._iptu > 0:
        partes.append(f"IPTU: R$ {devedor._iptu:.2f}")
    if not isinstance(devedor._cond, str) and devedor._cond is not None and devedor._cond > 0:
        partes.append(f"Taxa de condomínio: R$ {devedor._cond:.2f}")
    if not isinstance(devedor._taxa, str) and devedor._cond is not None and devedor._cond > 0:
        partes.append(f"Taxa de administração: R$ {devedor._taxa:.2f}")
    if not isinstance(devedor._tarifa, str) and devedor._cond is not None and devedor._cond > 0:
        partes.append(f"Tarifa de transferência: R$ {devedor._tarifa:.2f}")    
    partes.append(f"<br><b>Valor do Repasse: R$ {devedor._repasse:.2f}</b>")
    return "<br>".join(partes)

def extrair_mensagem_assunto(caminho_arquivo):
    """Extrai mensagem para o assunto do e-mail a partir do nome do arquivo."""
    if isinstance(caminho_arquivo, list):
        caminho_arquivo = caminho_arquivo[0]
    
    nome_arquivo = Path(caminho_arquivo).stem
    partes = nome_arquivo.split(" - ", 1)[1]
    return partes.rsplit(" - ", 1)[0] if " - " in partes else partes

def criar_email(devedor, texto_formatado, anexos, imagem="C:/Users/victo/Desktop/auto/Auto_emails_Vigor/lg.png"):
    """Cria e configura mensagem de e-mail, com opção de imagem embutida."""
    print(devedor._email_inquilino)
    if not devedor._endereco or not devedor._nome_inquilino or not devedor._email_inquilino:
        raise ValueError("Imóvel inválido: endereço, nome ou email não fornecidos.")
    msg = MIMEMultipart('related')
    msg['Subject'] = extrair_mensagem_assunto(anexos[0])
    msg['From'] = EMAIL_TO_PROP
    msg['To'] = EMAIL_TO_PROP

    # Corpo do email (com imagem embutida se fornecida)
    if devedor._cond is None:
        corpo_email = f"""
        <p></p>
        Boa tarde, {devedor._nome_inquilino.split()[0].capitalize()}<br>
        Segue o boleto referente ao  aluguel do mês de {MES_ANTERIOR} e taxa de condomínio referente ao mês vigente em anexo<br>
        <p><b>Valores:</b><br></p>
        {texto_formatado}  
        """
    else:
        corpo_email = f"""
        <p></p>
        Boa tarde, {devedor._nome_inquilino.split()[0].capitalize()}<br>
        Segue o boleto referente ao mês de {MES_ANTERIOR} em anexo<br>
        <p><b>Valores:</b><br></p>
        {texto_formatado}  
        """

    if imagem:
        # Adiciona a tag <img> ao corpo do email
        corpo_email += '<br><img src="cid:imagem1"><br>'

    msg_alt = MIMEMultipart('alternative')
    msg.attach(msg_alt)
    msg_alt.attach(MIMEText(corpo_email, 'html'))

    # Anexa imagem embutida, se fornecida
    if imagem:
        with open(imagem, 'rb') as img:
            mime_img = MIMEImage(img.read())
            mime_img.add_header('Content-ID', '<imagem1>')
            mime_img.add_header('Content-Disposition', 'inline', filename=Path(imagem).name)
            msg.attach(mime_img)

    # Anexa arquivos PDF
    for anexo in anexos:
        if anexo:
            with open(anexo, 'rb') as f:
                part = MIMEApplication(f.read(), _subtype='pdf')
                part.add_header('Content-Disposition', 'attachment', filename=Path(anexo).name)
                msg.attach(part)

    return msg

def enviar_email(msg):
    """Envia e-mail configurado."""
    with smtplib.SMTP('smtp.gmail.com', 587) as s:
        s.starttls()
        s.login(msg['From'], EMAIL_PASSWORD)
        s.send_message(msg)
    print(f'Email enviado para {msg["To"]}')

def enviar_email_proprietario(devedor):
    """Envia e-mail para proprietário."""
    if not devedor._pdfrepasse:
        return
    
    msg = criar_email(
        devedor,
        formatar_texto_proprietario(devedor),
        [devedor._pdfrepasse] if isinstance(devedor._pdfrepasse, str) else devedor._pdfrepasse
    )
    enviar_email(msg)

def enviar_email_inquilino(devedor):
    """Envia e-mail para inquilino."""
    if not devedor._pdfboleto:
        return
    
    anexos = [devedor._pdfboleto]
    if devedor._pdfcond:
        anexos.append(devedor._pdfcond)
    
    msg = criar_email(
        devedor,
        formatar_texto_inquilino(devedor),
        anexos
    )
    enviar_email(msg)

def main():
    """Função principal."""
    devedores = ler_planilha(ARQ_EXCEL, QTD)

    app = EmailInterface(devedores)
    app.run()

if __name__ == "__main__":
    main()